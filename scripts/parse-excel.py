#!/usr/bin/env python3
"""
parse-excel.py — Extract data from TableX's Excel files to JSON for the dashboard.

Source: /Users/dannybreckenridge/Documents/Clear ph/Clients/TableX/Claude/Quoting Docs and Sheets/
Output: /Users/dannybreckenridge/Applications/tablex-quoting-dashboard/src/data/

Outputs:
  1. profit-analysis.json  — ProfitAnalysis sheet from Quote Table (~680 valid rows)
  2. product-catalog.json   — TableX sheet from Quote Table (~6,100 valid rows)
  3. quote-queue.json       — Quote Queue (~3,635 data rows)
  4. staff.json             — Staff list from Quote Template
  5. dealers.json           — Dealer list from Quote Template
"""

import json
import os
import re
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime

import openpyxl

# ── Paths ────────────────────────────────────────────────────────────────────

SRC_DIR = "/Users/dannybreckenridge/Documents/Clear ph/Clients/TableX/Claude/Quoting Docs and Sheets"
OUT_DIR = "/Users/dannybreckenridge/Applications/tablex-quoting-dashboard/src/data"

QUOTE_TABLE_PATH = os.path.join(SRC_DIR, "2026 TableX Quote Table.xlsx")
QUOTE_QUEUE_PATH = os.path.join(SRC_DIR, "2026 QUOTE QUEUE.xlsx")
QUOTE_TEMPLATE_PATH = os.path.join(SRC_DIR, "MAF-2026 Quote Template.xlsx")

os.makedirs(OUT_DIR, exist_ok=True)


def write_json(data, filename):
    """Write data to a JSON file in the output directory."""
    path = os.path.join(OUT_DIR, filename)
    with open(path, "w") as f:
        json.dump(data, f, indent=2, default=str)
    print(f"  -> Wrote {filename}: {len(data)} records")


def safe_float(val, default=None):
    """Convert a value to float, returning default if not possible."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_str(val):
    """Convert a value to string, returning empty string for None."""
    if val is None:
        return ""
    return str(val).strip()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 1. PROFIT ANALYSIS — from Quote Table "ProfitAnalysis" sheet
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def parse_profit_analysis():
    """
    ProfitAnalysis sheet column mapping (from row 1 headers):
      A: Tag/label    B: Qty           C: Product (SKU)
      D: Top cost     E: Route cost    F: Bases cost
      G: NE/FD/FT     H: ASB/GN/LC     I: ASB/GN/LC (2nd)
      J: Smith/Valley (subtotal)       K: LF
      L: (unlabeled)  M: Frt In        N: PKG
      O: Total Cost   P: Frt Out %     Q: (cost + frt out, calculated)
      R: GPM          S: Comm.         T: Standard (selling price)
      U: Net Profit   V: List Price    W: Disc.
      X: Net Price    Y: new net profit Z: Notes

    Valid data rows have a SKU in column C (alphanumeric pattern with digits).
    Rows 681+ are formula residuals (no SKU, just PKG=15).
    """
    print("\n[1/5] Parsing ProfitAnalysis sheet...")
    wb = openpyxl.load_workbook(QUOTE_TABLE_PATH, read_only=True, data_only=True)
    ws = wb["ProfitAnalysis"]

    records = []
    for row in ws.iter_rows(min_row=2, max_row=1029, max_col=26, values_only=False):
        vals = {}
        for c in row:
            try:
                vals[c.column] = c.value
            except AttributeError:
                continue  # Skip EmptyCell objects in read_only mode

        sku = safe_str(vals.get(3))  # C = Product/SKU
        # Skip rows without a valid SKU
        if not sku or not any(ch.isdigit() for ch in sku):
            continue

        # Extract series code from SKU: first 2 digits (or "SP-XX" prefix)
        series_match = re.match(r"(?:SP-)?(\d{2})", sku)
        series = series_match.group(1) if series_match else ""

        # Discount tier from col W — 0.68 = 50/20/20, 0.66, 0.64, 0.62, 0.60 = 50/20
        disc = safe_float(vals.get(23))  # W
        list_price = safe_float(vals.get(22), 0)  # V

        # Calculate prices at each discount tier from list price
        # Cascading: 50/20 = list * 0.50 * 0.80 = 0.40
        # 50/20/5 = 0.40 * 0.95 = 0.38
        # 50/20/10 = 0.40 * 0.90 = 0.36
        # 50/20/15 = 0.40 * 0.85 = 0.34
        # 50/20/20 = 0.40 * 0.80 = 0.32
        price_50_20 = round(list_price * 0.40, 2) if list_price else None
        price_50_20_5 = round(list_price * 0.38, 2) if list_price else None
        price_50_20_10 = round(list_price * 0.36, 2) if list_price else None
        price_50_20_15 = round(list_price * 0.34, 2) if list_price else None
        price_50_20_20 = round(list_price * 0.32, 2) if list_price else None

        record = {
            "tag": safe_str(vals.get(1)),           # A
            "qty": safe_float(vals.get(2)),          # B
            "sku": sku,                              # C
            "series": series,
            "topCost": safe_float(vals.get(4)),      # D
            "routeCost": safe_float(vals.get(5)),    # E
            "baseCost": safe_float(vals.get(6)),     # F
            "nestFoldCost": safe_float(vals.get(7)), # G (NE, FD, or FT)
            "asbGnLcCost1": safe_float(vals.get(8)), # H
            "asbGnLcCost2": safe_float(vals.get(9)), # I
            "assemblyCost": safe_float(vals.get(10)), # J (Smith/Valley subtotal)
            "lfCost": safe_float(vals.get(11)),       # K
            "freightInCost": safe_float(vals.get(13)), # M
            "packagingCost": safe_float(vals.get(14)), # N
            "totalCost": safe_float(vals.get(15)),    # O
            "freightOutPct": safe_float(vals.get(16)), # P
            "gpm": safe_float(vals.get(18)),          # R
            "commission": safe_float(vals.get(19)),   # S
            "standardPrice": safe_float(vals.get(20)), # T
            "netProfit": safe_float(vals.get(21)),    # U
            "listPrice": safe_float(vals.get(22)),    # V
            "discountFactor": disc,                    # W
            "netPrice": safe_float(vals.get(24)),     # X
            "newNetProfit": safe_float(vals.get(25)), # Y
            "notes": safe_str(vals.get(26)),          # Z
            "price_50_20": price_50_20,
            "price_50_20_5": price_50_20_5,
            "price_50_20_10": price_50_20_10,
            "price_50_20_15": price_50_20_15,
            "price_50_20_20": price_50_20_20,
        }
        records.append(record)

    wb.close()
    write_json(records, "profit-analysis.json")
    return len(records)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 2. PRODUCT CATALOG — from Quote Table "TableX" sheet
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def decode_sku(sku):
    """
    Parse a TableX SKU into components.
    Example: 99SQ3030QD16-3P → series=99, shape=SQ, size=3030, base=QD16, posts=3P
    Patterns:
      - Optional SP- prefix (special)
      - 2-digit series
      - 2-letter shape code (TC, SQ, RD, SC, CS, HR, RT, etc.)
      - 4-digit size (W x D in inches, e.g. 3048 = 30x48)
      - Base code (T, U, X, QD, D + size)
      - Suffixes: -3P, -LC, -SH.XX, -NE, -FD, -CH, -GR.A, -PRE, -OS, etc.
    """
    shape_codes = {
        "TC": "Rectangular",
        "SQ": "Square",
        "RD": "Round",
        "SC": "Semicircle",
        "CS": "C-Shape",
        "HR": "Horse Race",
        "RT": "Racetrack",
        "OV": "Oval",
        "KD": "Kidney",
        "TR": "Trapezoid",
        "PW": "Power",
    }

    result = {"shape": "", "shapeName": "", "size": "", "baseType": ""}

    # Strip SP- prefix
    clean = re.sub(r"^SP-", "", sku)

    # Match: series (2 digits) + shape (2 letters) + size (4+ digits)
    m = re.match(r"(\d{2})([A-Z]{2})(\d{4,})", clean)
    if m:
        result["shape"] = m.group(2)
        result["shapeName"] = shape_codes.get(m.group(2), m.group(2))

        size_str = m.group(3)
        # Parse size: first 2 digits x next 2 digits (e.g., 3048 → 30"x48")
        if len(size_str) >= 4:
            w = size_str[:2]
            d = size_str[2:4]
            result["size"] = f'{w}"x{d}"'
        else:
            result["size"] = size_str

        # Base type follows size
        remainder = clean[m.end():]
        base_match = re.match(r"([A-Z]+\d*)", remainder)
        if base_match:
            result["baseType"] = base_match.group(1)

    return result


def parse_product_catalog():
    """
    TableX sheet column mapping (from row 1 headers):
      A: DATE         B: (label/notes) C: Qty
      D: Product (SKU)                 E: Top cost
      F: Route cost   G: Bases per table
      H: NE/FD/FT     I: ASB/GN/LC     J: ASB/GN/LC (2nd)
      K: Smith/Valley  L: LF            M: Edge
      N: Frt In        O: PKG           P: Total Cost
      Q: Frt Out %     R: (cost+frt)    S: GPM
      T: Comm.         U: Standard      V: Net Profit
      W: List Price    X: Disc.         Y: Net Price
      Z: new net profit
      AA: SKID 96"
      AD: Notes

    Valid rows: column D contains a SKU-like string (alphanumeric with digits, len > 5).
    Rows after ~8734 are formula residuals with no SKU.
    """
    print("\n[2/5] Parsing TableX product catalog sheet...")
    wb = openpyxl.load_workbook(QUOTE_TABLE_PATH, read_only=True, data_only=True)
    ws = wb["TableX"]

    records = []
    for row in ws.iter_rows(min_row=2, max_row=9068, max_col=31, values_only=False):
        vals = {}
        for c in row:
            try:
                vals[c.column] = c.value
            except AttributeError:
                continue  # Skip EmptyCell objects in read_only mode

        sku = safe_str(vals.get(4))  # D = Product/SKU
        # Filter: must have digits, be long enough, and look like a SKU
        if not sku or len(sku) < 6 or not any(ch.isdigit() for ch in sku):
            continue
        # Skip description-only rows (e.g., "5% price increase + 7% premium laminate")
        if sku.startswith("5%") or sku.startswith("10%") or "price" in sku.lower():
            continue

        # Extract series
        clean_sku = re.sub(r"^SP-", "", sku)
        series_match = re.match(r"(\d{2})", clean_sku)
        series = series_match.group(1) if series_match else ""

        # Decode SKU components
        sku_parts = decode_sku(sku)

        list_price = safe_float(vals.get(23), 0)  # W = List Price

        # Calculate discount tier prices from list
        price_50_20 = round(list_price * 0.40, 2) if list_price else None
        price_50_20_5 = round(list_price * 0.38, 2) if list_price else None
        price_50_20_10 = round(list_price * 0.36, 2) if list_price else None
        price_50_20_15 = round(list_price * 0.34, 2) if list_price else None
        price_50_20_20 = round(list_price * 0.32, 2) if list_price else None

        record = {
            "sku": sku,
            "series": series,
            "shape": sku_parts["shape"],
            "shapeName": sku_parts["shapeName"],
            "size": sku_parts["size"],
            "baseType": sku_parts["baseType"],
            "topCost": safe_float(vals.get(5)),       # E
            "routeCost": safe_float(vals.get(6)),      # F
            "baseCost": safe_float(vals.get(7)),       # G
            "nestFoldCost": safe_float(vals.get(8)),   # H
            "asbGnLcCost1": safe_float(vals.get(9)),   # I
            "asbGnLcCost2": safe_float(vals.get(10)),  # J
            "assemblyCost": safe_float(vals.get(11)),  # K (Smith/Valley)
            "lfCost": safe_float(vals.get(12)),        # L
            "edgeCost": safe_float(vals.get(13)),      # M
            "freightInCost": safe_float(vals.get(14)), # N
            "packagingCost": safe_float(vals.get(15)), # O
            "totalCost": safe_float(vals.get(16)),     # P
            "freightOutPct": safe_float(vals.get(17)), # Q
            "gpm": safe_float(vals.get(19)),           # S
            "commission": safe_float(vals.get(20)),    # T
            "standardPrice": safe_float(vals.get(21)), # U
            "netProfit": safe_float(vals.get(22)),     # V
            "listPrice": safe_float(vals.get(23)),     # W
            "discountFactor": safe_float(vals.get(24)), # X
            "netPrice": safe_float(vals.get(25)),      # Y
            "newNetProfit": safe_float(vals.get(26)),  # Z
            "price_50_20": price_50_20,
            "price_50_20_5": price_50_20_5,
            "price_50_20_10": price_50_20_10,
            "price_50_20_15": price_50_20_15,
            "price_50_20_20": price_50_20_20,
            "notes": safe_str(vals.get(30)),           # AD
        }
        records.append(record)

    wb.close()
    write_json(records, "product-catalog.json")
    return len(records)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 3. QUOTE QUEUE — parsed via raw XML to bypass stylesheet bug
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def normalize_date(raw):
    """
    Attempt to parse dates like "Feb 10 / 8:48pm" into ISO 8601 strings.
    Returns the raw string if parsing fails.
    """
    if not raw or not isinstance(raw, str):
        return raw

    raw = raw.strip()

    # Common patterns: "Feb 10 / 8:48pm", "July 11 / 10:04am", "Sept 23 / 1:38pm"
    # Month names can be 3-letter abbreviations or longer (June, July, Sept, etc.)
    # Also handles Excel serial date numbers (e.g., "45037")
    patterns = [
        # "Feb 10 / 8:48pm" or "July 11 / 10:04am"
        (r"(\w{3,9})\s+(\d{1,2})\s*/\s*(\d{1,2}):(\d{2})\s*(am|pm|AM|PM)",
         lambda m: _build_dt(m.group(1), m.group(2), m.group(3), m.group(4), m.group(5))),
        # "Feb 10 / 8:48 pm" (space before am/pm)
        (r"(\w{3,9})\s+(\d{1,2})\s*/\s*(\d{1,2}):(\d{2})\s+(am|pm|AM|PM)",
         lambda m: _build_dt(m.group(1), m.group(2), m.group(3), m.group(4), m.group(5))),
        # "Feb 10 / 8pm"  (no minutes)
        (r"(\w{3,9})\s+(\d{1,2})\s*/\s*(\d{1,2})\s*(am|pm|AM|PM)",
         lambda m: _build_dt(m.group(1), m.group(2), m.group(3), "00", m.group(4))),
    ]

    for pattern, builder in patterns:
        m = re.match(pattern, raw)
        if m:
            try:
                return builder(m)
            except (ValueError, KeyError):
                pass

    # Handle Excel serial date numbers (e.g., "45037" = days since 1899-12-30)
    if raw.isdigit() and 40000 < int(raw) < 50000:
        try:
            from datetime import timedelta
            base = datetime(1899, 12, 30)
            dt = base + timedelta(days=int(raw))
            return dt.strftime("%m-%dT00:00")
        except (ValueError, OverflowError):
            pass

    return raw  # Return raw string if no pattern matches


def _build_dt(month_str, day_str, hour_str, min_str, ampm):
    """Build an ISO date string from components. Year is not in the data."""
    month_map = {
        "jan": 1, "january": 1,
        "feb": 2, "february": 2,
        "mar": 3, "march": 3,
        "apr": 4, "april": 4,
        "may": 5,
        "jun": 6, "june": 6,
        "jul": 7, "july": 7,
        "aug": 8, "august": 8,
        "sep": 9, "sept": 9, "september": 9,
        "oct": 10, "october": 10,
        "nov": 11, "november": 11,
        "dec": 12, "december": 12,
    }
    month = month_map.get(month_str.lower())
    if month is None:
        raise ValueError(f"Unknown month: {month_str}")

    day = int(day_str)
    hour = int(hour_str)
    minute = int(min_str)

    # Convert 12-hour to 24-hour
    ampm_lower = ampm.lower()
    if ampm_lower == "pm" and hour != 12:
        hour += 12
    elif ampm_lower == "am" and hour == 12:
        hour = 0

    # We don't have the year in the date string, so we'll leave it out
    # and just return a month-day-time string. The year can be inferred
    # from position in the queue (2023-2026).
    return f"{month:02d}-{day:02d}T{hour:02d}:{minute:02d}"


def parse_quote_queue():
    """
    Parse Quote Queue via raw XML (zipfile) to bypass stylesheet issues.

    Column mapping (from row 2 headers):
      A: EMAIL FROM
      B: DATE / TIME REC'D
      C: QUOTE # / EMAIL QUOTE / SO #
      D: DEALER / PROJECT NAME
      E: SPECIAL?
      F: MAF, SS, or MM?
      G: STATUS / COMPLETED

    Row 1 = title, Row 2 = headers, Rows 3-4 = section headers.
    Data starts at row 5.
    Some rows are section headers (e.g., "COMPLETED ALL 2023...").
    """
    print("\n[3/5] Parsing Quote Queue via raw XML...")

    ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

    with zipfile.ZipFile(QUOTE_QUEUE_PATH, "r") as z:
        # Load shared strings
        with z.open("xl/sharedStrings.xml") as f:
            tree = ET.parse(f)
            root = tree.getroot()
            strings = []
            for si in root.findall(f".//{{{ns}}}si"):
                # Handle rich text (multiple <r><t> elements)
                texts = []
                for t_el in si.findall(f".//{{{ns}}}t"):
                    if t_el.text:
                        texts.append(t_el.text)
                strings.append("".join(texts))

        # Parse sheet data
        with z.open("xl/worksheets/sheet1.xml") as f:
            tree = ET.parse(f)
            root = tree.getroot()
            xml_rows = root.findall(f".//{{{ns}}}row")

        # Column letter to index: A=0, B=1, ..., G=6
        def col_index(ref):
            """Extract column letter(s) from cell reference like 'A5' and return 0-based index."""
            letters = re.match(r"([A-Z]+)", ref).group(1)
            idx = 0
            for ch in letters:
                idx = idx * 26 + (ord(ch) - ord("A") + 1)
            return idx - 1  # 0-based

        records = []
        # Track current year section for date inference
        current_year = 2023

        for xml_row in xml_rows:
            row_num = int(xml_row.attrib.get("r", 0))
            if row_num <= 2:
                continue  # Skip title and header rows

            # Extract cell values
            cells = {}
            for c_el in xml_row.findall(f"{{{ns}}}c"):
                ref = c_el.attrib.get("r", "")
                t = c_el.attrib.get("t", "")
                v_el = c_el.find(f"{{{ns}}}v")
                val = None
                if v_el is not None and v_el.text is not None:
                    if t == "s":
                        val = strings[int(v_el.text)]
                    else:
                        val = v_el.text
                ci = col_index(ref)
                cells[ci] = val

            a_val = safe_str(cells.get(0))  # A: Email From
            b_val = safe_str(cells.get(1))  # B: Date/Time
            c_val = safe_str(cells.get(2))  # C: Quote #
            d_val = safe_str(cells.get(3))  # D: Dealer/Project
            e_val = safe_str(cells.get(4))  # E: Special?
            f_val = safe_str(cells.get(5))  # F: Staff
            g_val = safe_str(cells.get(6))  # G: Status

            # Skip section header rows (they span across merged cells with long descriptions)
            if a_val and ("COMPLETED" in a_val.upper() or "QUEUE" in a_val.upper()):
                # Try to extract year from section header
                year_match = re.search(r"20(2[3-9])", a_val)
                if year_match:
                    current_year = int("20" + year_match.group(1))
                continue

            # Skip rows with no meaningful data (no email from AND no dealer)
            if not a_val and not d_val:
                continue

            # Normalize Special? to boolean
            special = None
            if e_val:
                e_upper = e_val.upper().strip()
                if e_upper in ("YES", "Y"):
                    special = True
                elif e_upper in ("NO", "N", "-"):
                    special = False

            # Normalize date
            date_normalized = normalize_date(b_val)
            status_normalized = normalize_date(g_val)

            record = {
                "rowNum": row_num,
                "emailFrom": a_val,
                "dateTime": b_val,
                "dateNormalized": date_normalized,
                "year": current_year,
                "quoteNumber": c_val,
                "dealerProject": d_val,
                "special": special,
                "staff": f_val.strip(),
                "status": g_val,
                "statusNormalized": status_normalized,
            }
            records.append(record)

        # Second pass: infer year from quote numbers where available
        # Quote numbers like 23.MAF.xxx → year 2023, 24.xxx → 2024, etc.
        for rec in records:
            qn = rec["quoteNumber"]
            year_from_qn = re.match(r"^(2[3-9])\.", qn)
            if year_from_qn:
                rec["year"] = 2000 + int(year_from_qn.group(1))

    write_json(records, "quote-queue.json")
    return len(records)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 4 & 5. STAFF + DEALERS — from Quote Template "Dropdown Menus" sheet
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def parse_template_dropdowns():
    """
    Dropdown Menus sheet layout:
      Rows 1-4: Staff (col A = name, col B = email). Row 5-6 are blank.
      Rows 7-19: Dealers (col A = dealer name).

    The split is: staff have an email in column B; dealers don't
    (except row 11 CRG has "Data Validation" in B which is a label, not an email).
    """
    print("\n[4/5] Parsing staff list from Quote Template...")
    print("[5/5] Parsing dealer list from Quote Template...")

    wb = openpyxl.load_workbook(QUOTE_TEMPLATE_PATH, read_only=True, data_only=True)
    ws = wb["Dropdown Menus"]

    staff = []
    dealers = []

    for row in ws.iter_rows(min_row=1, max_row=50, max_col=3, values_only=False):
        vals = {}
        for c in row:
            try:
                vals[c.column_letter] = c.value
            except AttributeError:
                continue  # Skip EmptyCell objects in read_only mode
        name = safe_str(vals.get("A"))
        email = safe_str(vals.get("B"))

        if not name:
            continue

        # Staff have a proper email address
        if "@" in email:
            staff.append({
                "name": name,
                "email": email,
                "initials": _get_initials(name),
            })
        else:
            # It's a dealer (skip non-dealer labels)
            if email and "validation" in email.lower():
                pass  # skip metadata label
            dealers.append({
                "name": name,
            })

    wb.close()

    write_json(staff, "staff.json")
    write_json(dealers, "dealers.json")
    return len(staff), len(dealers)


def _get_initials(name):
    """Generate initials from a full name: 'Mark Fleck' → 'MAF'."""
    # Known mappings from the quote numbering convention
    known = {
        "Mark Fleck": "MAF",
        "Brian Craig": "BC",
        "Samatha Sander": "SS",
        "Maya Mitchell": "MM",
    }
    if name in known:
        return known[name]
    parts = name.split()
    return "".join(p[0].upper() for p in parts if p)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# MAIN
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

if __name__ == "__main__":
    print("=" * 60)
    print("TableX Excel → JSON Data Extraction")
    print("=" * 60)

    n_profit = parse_profit_analysis()
    n_catalog = parse_product_catalog()
    n_queue = parse_quote_queue()
    n_staff, n_dealers = parse_template_dropdowns()

    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"  profit-analysis.json : {n_profit:,} rows")
    print(f"  product-catalog.json : {n_catalog:,} rows")
    print(f"  quote-queue.json     : {n_queue:,} rows")
    print(f"  staff.json           : {n_staff} staff members")
    print(f"  dealers.json         : {n_dealers} dealers")
    print(f"\nAll files written to: {OUT_DIR}")
    print("=" * 60)
